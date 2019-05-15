# -*- coding: utf-8 -*-
"""
Created on Tue May 14 11:07:55 2019

JTT Parsing, output Do Not Pay and AFR Analysis 

### COMPLETED - NEEDS INTEGRATION ###  Read in the JTT file, output a file with the Do Not Pay, and AFR analysis. 
### COMPLETED - NEEDS INTEGRATION ###  Open Excel File
# If loop through each row, and if the AG == "Y" && AI == "Y" then add that row's columns to Do Pay List.xlsx
# Open the Do Not Pay central file in /TAXES/TAX_TOOLS/Do_Not_Pay.xlsx and 
# Add the Do Not Pay entry 
# Save the Do Not Pay workbook
# Close the Do Not Pay File
# If loop through each row, and if the AG == "N" && AI != "N" then add that row's columns 38, 39, 40, 41 to a Do_Not_Pay.xlsx
# Open the AFR analysis central file in Shared/Escrow_Administration
# Add the AFR analysis entry
# Save the AFR Analysis workbook
# Close the AFR Analysis workbook
# If loop through each row, and if the AI column.value == "Y" then add that row's columns 37, 38, 39 to a AFR_analysis_sheet.xlsx
# Create an Excel file for the Do_Pay_List.xlsx 

@author: austin.schrader
"""
# Copy the workbook
import pandas
# Actually access the workbook
from openpyxl import load_workbook

# Job fileName and job stateCode, hardcoded which should be replaced later
fileName = "\\NJ66666666666666666"
stateCode = "\\NJ"

# Establishes the path for the .xlsx file (the version that we can work with)
jobPath = ("\\\\cottonwood\\Users\\Shared\\Taxes\\CTA Paid Taxes\\2019" + stateCode + fileName)
exportFilePath = "\\\\cottonwood\\Users\\Shared\\Taxes\\CTA Paid Taxes\\2019" + stateCode + fileName + "\\For Russell to review-Loan Template.xlsm"

# Reads the excel and parses out columns 0,1,4,6 etc (all the ones we need.)
# Then, it exports the file to the jobPath + \\output.xlsx
dataframe = pandas.read_excel(exportFilePath, parse_cols = [0,1,2,3,4,23,32,34,36])
dataframe.to_excel(jobPath + "\\managementoutput.xlsx")

# Creates a specific dataframe that contains the records of loans that need management approval. This can be added
# to the do not pay, seek management approval section of the Do_Not_Pay.xlsx spreadsheet next
mgmtDataFrame = dataframe[dataframe["SentToMgmtFlag"] == "Y"]
mgmtDataFrame.to_excel(jobPath + "\\addtomgmt.xlsx", header=False) #, header=False, index=False
#print(mgmtDataFrame)

# Creates a specific dataframe that contains the records of loans that will not be paid today. The reason include delinquency on the loan
# In which case, they should be deleted from the Export-TCS36501 file
payDataFrame = dataframe[dataframe["PayFlag"] == "N"]
payDataFrame.to_excel(jobPath + "\\addtodonotpay.xlsx", header=False) #, header=False, index=False


#print(payDataFrame)

# =============================================================================
# book = load_workbook("dontpaythat.xlsx")
# writer = pandas.ExcelWriter("dontpaythat.xlsx", engine="openpyxl")
# writer.book = book
# writer.sheets = {ws.title: ws for ws in book.worksheets}
# =============================================================================

#for sheetname in writer.sheets:
#    payDataFrame.to_excel(writer, sheet_name="Sheet1", startrow=writer.sheets["Sheet1"].max_row, index == False)
                 


# Creates a specific dataframe that contains the records of loans that need EA analysis. Thus, these loans should be 
# added to the EA Analysis spreadsheet
eaDataFrame = dataframe[dataframe["EAFlag"] == "Y"]
eaDataFrame.to_excel(jobPath + "\\addtoea.xlsx", header=False)
#print(eaDataFrame)



# =============================================================================
# dataframe.to_excel(jobPath + "\\donotpay.xlsx")
# dataframe.to_excel(jobPath + "\\addtoafr.xlsx")
# dataframe.to_excel(jobPath + "\\askmgmt.xlsx")
# =============================================================================

#do not pay = \\cottonwood\Users\Shared\Taxes\TAX_TOOLS




# Read in the JTT file, output a file with the Do Not Pay, and AFR analysis. 
# Open Excel File
# If loop through each row, and if the AG == "Y" && AI == "Y" then add that row's columns to Do Pay List.xlsx
# Open the Do Not Pay central file in /TAXES/TAX_TOOLS/Do_Not_Pay.xlsx and 
# Add the Do Not Pay entry 
# Save the Do Not Pay workbook
# Close the Do Not Pay File
# If loop through each row, and if the AG == "N" && AI != "N" then add that row's columns 38, 39, 40, 41 to a Do_Not_Pay.xlsx
# Open the AFR analysis central file in Shared/Escrow_Administration
# Add the AFR analysis entry
# Save the AFR Analysis workbook
# Close the AFR Analysis workbook
# If loop through each row, and if the AI column.value == "Y" then add that row's columns 37, 38, 39 to a AFR_analysis_sheet.xlsx
# Create an Excel file for the Do_Pay_List.xlsx 

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False, 
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    from openpyxl import load_workbook

    import pandas as pd

    print("We're here!1")
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl')

    # Python 2.x: define [FileNotFoundError] exception if it doesn't exist 
    try:
        FileNotFoundError
    except NameError:
        FileNotFoundError = IOError


    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
        print("We're here!2")
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
            print("We're here! 3")
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
            print("We're here!4")
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0
        print("We're here! 5")

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()

print(payDataFrame)

append_df_to_excel("\\\\cottonwood\\Users\\Shared\\Taxes\\CTA Paid Taxes\\2019\\NJ\\NJ66666666666666666\\Do not Pay.xlsx", payDataFrame, sheet_name="LTR111")
#append_df_to_excel("\\\\cottonwood\\Users\\Shared\\Taxes\\CTA Paid Taxes\\2019\\NJ\\NJ66666666666666666\\AFR Escrow Analysis.xlsx", eaDataFrame, sheet_name="Tax", header=False)
#append_df_to_excel("\\\\cottonwood\\Users\\Shared\\Taxes\\CTA Paid Taxes\\2019\\NJ\\NJ66666666666666666\\Do not Pay.xlsx", payDataFrame, sheet_name="LTR111", header=False)